from openpyxl import load_workbook
import os

banned_file = "banned_list.xlsx"

def read_banned_list():
    if not os.path.exists(banned_file):
        return []
    wb = load_workbook(banned_file)
    ws = wb.active
    banned_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(row, tuple) and len(row) >= 2 and row[0] is not None:
            banned_list.append((row[0], row[1]))  # Read name and banned until date
    return banned_list

def view_banned_list():
    banned_list = read_banned_list()
    print("Banned List:")
    for person, banned_until in banned_list:
        print(f'{person} - Banned until {banned_until}')