from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

log_file = "id_check_log.xlsx"

def log_id_check(name, status, dob=None, id_type=None):
    today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = load_workbook(log_file)
    ws = wb.active
    ws.append([name, today, status, dob, id_type])
    wb.save(log_file)

def display_logs():
    if not os.path.exists(log_file):
        print("No logs found.")
        return
    wb = load_workbook(log_file)
    ws = wb.active
    print("ID Check Logs:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)