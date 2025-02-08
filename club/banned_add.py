from openpyxl import load_workbook, Workbook
import os

banned_file = "banned_list.xlsx"

def read_banned_list():
    if not os.path.exists(banned_file):
        return []
    wb = load_workbook(banned_file)
    ws = wb.active
    banned_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(row, tuple) and len(row) >= 2 and row[0] is not None:
            banned_list.append((row[0], row[1]))  # Read name and banned until date
    return banned_list

def write_banned_list(banned_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Banned"
    ws.append(["Name", "Banned Until"])  # Adding headers
    for person, banned_until in banned_list:
        ws.append([person, banned_until])
    wb.save(banned_file)

def add_to_banned_list():
    banned_list = read_banned_list()
    name = input('Enter the name of the person to ban: ')
    banned_until = input('Enter the date until they are banned (dd/mm/yyyy): ')
    if name not in [person for person, _ in banned_list]:
        banned_list.append((name, banned_until))
        write_banned_list(banned_list)
        print(f'{name} has been added to the banned list until {banned_until}.')
    else:
        print(f'{name} is already on the banned list.')