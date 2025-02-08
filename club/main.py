from datetime import datetime, timedelta
import sys
import os
from openpyxl import load_workbook, Workbook

password = 'Brookhouse01!'
log_file = "id_check_log.xlsx"
banned_file = "banned_list.xlsx"

# Check if the id_check_log.xlsx file exists, if not, create it
if not os.path.exists(log_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["Name", "Date", "Status", "Date of Birth", "ID Type"])  # Adding headers
    wb.save(log_file)

def log_id_check(name, status, dob=None, id_type=None):
    today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = load_workbook(log_file)
    ws = wb.active
    ws.append([name, today, status, dob, id_type])
    wb.save(log_file)

def read_banned_list():
    if not os.path.exists(banned_file):
        return []
    wb = load_workbook(banned_file)
    ws = wb.active
    banned_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(row, tuple) and len(row) >= 2 and row[0] is not None:
            banned_list.append((row[0], row[1]))  # Read name and banned until date
    return banned_list

def write_banned_list(banned_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Banned"
    ws.append(["Name", "Banned Until"])  # Adding headers
    for person, banned_until in banned_list:
        ws.append([person, banned_until])
    wb.save(banned_file)

def banned_from_club(name):
    """ This will check to see if a person is banned from the club"""
    banned_people = read_banned_list()
    for person, banned_until in banned_people:
        if name == person:
            print(f'Do not allow entry to the club!!! Banned until {banned_until}')
            log_id_check(name, 'Banned')
            return False
    print('Now lets check their age!')
    return True

def auth_id_type():
    return ['passport', 'driving licence', 'national id card']

def age_check(name):
    """ This will check to see if a person is old enough to enter the club
        and check their ID """
    how_old = int(input('How old is the person trying to enter? '))
    if how_old < 18:
        print('This person is not old enough to enter!!!')
        log_id_check(name, 'Underage')
        return False
    else:
        print('This person is old enough to enter. Let’s check their ID!')
        return True

def id_check(name):
    """ This will check their ID to see if it is valid! """
    valid_ids = auth_id_type()
    id_type = input('What form of ID are you presented with? ')
    if id_type not in valid_ids:
        print('Invalid ID type, please ask for another form of ID or decline entry!')
        log_id_check(name, 'Invalid ID')
        return False
    else:
        print('ID type is valid. Let’s check their date of birth!')
        return id_type

def is_legal_age(birthdate, legal_age=18):
    birthdate = datetime.strptime(birthdate, "%Y-%m-%d").date()
    legal_age_date = birthdate + timedelta(days=legal_age * 365.25)
    today = datetime.now().date()
    return today >= legal_age_date

def dob_check(name, id_type):
    day = int(input('What day of the month does the birthdate fall on? '))
    month = int(input('What month does their birthday fall on? '))
    year = int(input('What year were they born? '))
    birthdate = f"{year}-{month:02d}-{day:02d}"  # Format birthdate as YYYY-MM-DD
    if is_legal_age(birthdate):
        print('Allow Entry!!!')
        log_id_check(name, 'Allowed', birthdate, id_type)
    else:
        print('Do not allow entry!!!')
        log_id_check(name, 'Underage', birthdate, id_type)

def nextp():
    answer = 'y'
    while answer == 'y':
        name = input('What is the name of the person trying to enter the club? ')
        if not banned_from_club(name):
            answer = input("Press 'y' when you are ready to check the next person or press 'n' to leave the program: ")
            continue
        if not age_check(name):
            answer = input("Press 'y' when you are ready to check the next person or press 'n' to leave the program: ")
            continue
        id_type = id_check(name)
        if not id_type:
            answer = input("Press 'y' when you are ready to check the next person or press 'n' to leave the program: ")
            continue
        dob_check(name, id_type)
        answer = input("Press 'y' when you are ready to check the next person or press 'n' to leave the program: ")

def manage_banned_list():
    banned_list = read_banned_list()
    while True:
        action = input('Type "add" to add a person to the banned list, \nType "remove" to remove a person, \nType "view" to view the banned list, \nType "back" to return to the main menu: ')
        if action == 'add':
            name = input('Enter the name of the person to ban: ')
            banned_until = input('Enter the date until they are banned (dd/mm/yyyy): ')
            if name not in [person for person, _ in banned_list]:
                banned_list.append((name, banned_until))
                write_banned_list(banned_list)
                print(f'{name} has been added to the banned list until {banned_until}.')
            else:
                print(f'{name} is already on the banned list.')
        elif action == 'remove':
            name = input('Enter the name of the person to remove from the banned list: ')
            if name in [person for person, _ in banned_list]:
                banned_list = [(person, date) for person, date in banned_list if person != name]
                write_banned_list(banned_list)
                print(f'{name} has been removed from the banned list.')
            else:
                print(f'{name} is not on the banned list.')
        elif action == 'view':
            print("Banned List:")
            for person, banned_until in banned_list:
                print(f'{person} - Banned until {banned_until}')
        elif action == 'back':
            break
        else:
            print('Invalid choice, please try again.')

def display_logs():
    if not os.path.exists(log_file):
        print("No logs found.")
        return
    wb = load_workbook(log_file)
    ws = wb.active
    print("ID Check Logs:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)

def start_session():
    print("Session started.")
    main_menu()

def end_session():
    if password == input('Please enter the password to end the session: '):
        print("Session ended.")
        sys.exit()

def main_menu():
    while True:
        choice = input("Type 'banned' to manage banned list, 'check' to check IDs, 'logs' to view logs, 'exit' to end session: ")
        if choice == 'banned':
            manage_banned_list()
        elif choice == 'check':
            nextp()
        elif choice == 'logs':
            display_logs()
        elif choice == 'exit':
            end_session()
        else:
            print('Invalid choice, please try again.')

start_session()
