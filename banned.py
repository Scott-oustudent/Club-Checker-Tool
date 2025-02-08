# banned.py

from openpyxl import load_workbook, Workbook
import os

# Check if the banned_people.xlsx file exists, if not, create it
if not os.path.exists("banned_people.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Banned People"
    ws.append(["Name"])  # Adding header
    wb.save("banned_people.xlsx")

def load_banned_people():
    wb = load_workbook("banned_people.xlsx")
    ws = wb["Banned People"]
    return [row[0].value for row in ws.iter_rows(min_row=2, values_only=True)]

def save_banned_people(banned_people):
    wb = load_workbook("banned_people.xlsx")
    ws = wb["Banned People"]
    ws.delete_rows(2, ws.max_row)  # Clear existing data
    for person in banned_people:
        ws.append([person])
    wb.save("banned_people.xlsx")

banned_people = []

def banned_menu():
    print(f'The People who are banned from the club are: {banned_people}')
    bannedm = input("If you would like to add a person to the banned list, type 'add'. To remove someone from the banned list, type 'remove': ")
    if bannedm == 'add':
        bannedadd = input("What is the name of the person you would like to ban? ")
        banned_people.append(bannedadd)
        print(f'{bannedadd} has been added to the banned list')
        print(banned_people)
    elif bannedm == 'remove':
        bannedrm = input("What is the name of the person you would like to remove? ")
        if bannedrm in banned_people:
            banned_people.remove(bannedrm)
            print(f'{bannedrm} has been removed from the banned list!')
        else:
            print(f'{bannedrm} is not in the banned list!')
        print(banned_people)
    else:
        print('Invalid choice, returning to the main menu.')
